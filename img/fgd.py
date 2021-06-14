#! python3
import openpyxl,os,sys,shutil,time,subprocess,pyperclip,tkinter,random,docx,threading,webbrowser
from openpyxl.styles import Font
from tkinter import *
from PIL import ImageTk,Image
import requests
from lxml import etree
import lxml.html
import csv
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
put='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\10.Организации- участники соревнований'    
Komandi={'Локомотив-1':'lokomotiv.xlsx','Спартак':'spartakAkademia.xlsx','ЦСКА':'cska.xlsx','Динамо':'dinamo.xlsx','Чертаново':'chertanovo.xlsx',
         'ФШМ':'fshm.xlsx','Строгино':'strogino.xlsx','Химки':'himki.xlsx','Родина':'rodina.xlsx',
         'Спартак-2':'spartakMosKomSport.xlsx','ТорпедоМосКомСпорт':'torpedoMosKomSport.xlsx','Смена':'smena.xlsx','Сокол':'sokol.xlsx','Трудовые резервы':'trudovieRezervi.xlsx',
         'Крылья Советов':'krilyaSovetov.xlsx','Строгино-2':'strogino.xlsx','Приалит':'prialit.xlsx','Росич':'rosich.xlsx','ФШМ-2':'fshm.xlsx',
         'Спутник':'sputnik.xlsx','Чертаново-2':'chertanovo.xlsx','Буревестник':'burevestnik.xlsx','Кунцево':'kuncevo.xlsx','Динамо-2':'dinamo.xlsx','Красногвардеец':'krasnogvardeec.xlsx','Мегасфера':'megasfera.xlsx',
         'Савеловская':'savelovskaya.xlsx','Москвич':'moskvich.xlsx','СШ-4':'sh4.xlsx','САШ Олимпик':'sash80.xlsx','Торпедо-2':'torpedoMosKomSport.xlsx','Смена -2':'smena.xlsx','Молния':'molniya.xlsx',
         'Крылья Советов -2':'krilyaSovetov.xlsx','Савеловская-2':'savelovskaya.xlsx',
         'Воробьевы Горы':'vorobGory.xlsx','Лобня':'lobnya.xlsx','Крылья Советов-3':'krilyaSovetov.xlsx','Трудовые резервы-2':'trudovieRezervi.xlsx','СШ №101 Тушино':'tushino.xlsx','Царицыно':'tcaricino.xlsx','СШ №82':'sh82.xlsx','Троицк':'troick.xlsx','Красногвардеец-2':'krasnogvardeec.xlsx',
         'Митино':'mitino.xlsx','Витязь':'vityaz.xlsx','Спарта-Свиблово':'spartaSviblovo.xlsx','Смена-юниор':'smena.xlsx',
         'Авангард':'avangard.xlsx','Сахалинец':'sahalinec.xlsx','ОД-80':'od80.xlsx','МВА':'mva.xlsx',
         'Юнитра':'unitra.xlsx','Сокол-2':'sokol.xlsx','Мегасфера-2':'megasfera.xlsx','Лужники':'akademiaSalova.xlsx','Кунцево-2':'kuncevo.xlsx','ДЮСШ Торпедо':'torpedo.xlsx',
         'Будущее':'budugee.xlsx','Крепость':'krepost.xlsx','Савеловская-3':'','Луч':'','Адмирал':'','ФК АС':'',
         'Троицк':'troick.xlsx','СШ-4 Легион':'sh4.xlsx','Крылья Советов-Альянс':'','Росич-2':'rosich.xlsx','Центавр':'centavr.xlsx','ТорпедоЛФК':'','СШ-75 - 2':'','Спартак-2ЛФК':'','Трудовые резервы-2ЛФК':'trudovieRezervi.xlsx','СоколЛФК':'sokol.xlsx','Родина-м':'',
         'КрасногвардеецЛФК':'krasnogvardeec.xlsx' ,'КунцевоЛФК':'kuncevo.xlsx','Родина-2':'rodina.xlsx' ,'РосичЛФК':'rosich.xlsx','Сетунь-Кунцево':'setunKuncevo.xlsx','БуревестникЛФК':'burevestnik.xlsx',
         'Летний дождик':'letniyDogdik.xlsx','Зенит':'zenit.xlsx','ФШМ':'fshm.xlsx','СменаЛФК':'smena.xlsx','ГЕРАКЛИОН':'geraklion.xlsx','СШ-75':'','Зеленоград':'zelenograd.xlsx','Гефест':'gefest.xlsx','Трудовые резервыЛФК':'trudovieRezervi.xlsx',
         'ЛОКОМОТИВЛФК':'lokomotiv.xlsx','Строгино-м':'strogino.xlsx','Спортакадемклуб':''}# Луч,Адмирал, ФК АС, ЦЕНТАВР,СШ 75,
KlubnayaLigaSpisok=['Локомотив-1', 'Спартак', 'ЦСКА', 'Динамо', 'Чертаново', 'ФШМ', 'Строгино', 'Химки', 'Родина']
PervayaLigaLigaSpisok=['Спартак-2', 'ТорпедоМосКомСпорт', 'Смена', 'Сокол', 'Трудовые резервы', 'Крылья Советов', 'Строгино-2', 'Приалит', 'Росич', 'ФШМ-2', 'Спутник', 'Чертаново-2']
VtorayaLigaLigaSpisok=['Буревестник', 'Кунцево', 'Динамо-2', 'Красногвардеец', 'Мегасфера', 'Савеловская', 'Москвич', 'СШ-4', 'САШ Олимпик', 'Торпедо-2', 'Смена -2', 'Молния']
TretyaLigaLigaSpisok=['Крылья Советов -2', 'Савеловская-2', 'Воробьевы Горы', 'Крылья Советов-3', 'Трудовые резервы-2', 'СШ №101 Тушино', 'Царицыно', 'СШ №82', 'Троицк', 'Красногвардеец-2', 'Митино', 'Витязь', 'Спарта-Свиблово', 'Смена-юниор']
ChetvetrtayaLigaLigaSpisok=['Авангард', 'ОД-80', 'МВА', 'Юнитра', 'Сокол-2', 'Мегасфера-2', 'Лужники', 'Кунцево-2', 'ДЮСШ Торпедо', 'Будущее', 'Крепость', 'Савеловская-3', 'Луч', 'Адмирал', 'ФК АС']
DivizASpisok=['Родина-2', 'РосичЛФК', 'Сетунь-Кунцево', 'БуревестникЛФК', 'Летний дождик', 'Зенит', 'СменаЛФК', 'ГЕРАКЛИОН', 'СШ-75', 'Зеленоград', 'Гефест', 'Трудовые резервыЛФК', 'ЛОКОМОТИВЛФК', 'Строгино-м','Спортакадемклуб']
DicizBSpisok=['Троицк','СШ-4 Легион', 'Крылья Советов-Альянс', 'Росич-2', 'Центавр', 'ТорпедоЛФК', 'СШ-75 - 2', 'Спартак-2ЛФК', 'Трудовые резервы-2ЛФК', 'СоколЛФК', 'Родина-м', 'КрасногвардеецЛФК', 'КунцевоЛФК']
SpisokLeague=['Клубная Лига','Первая Лига','Вторая Лига','Третья Лига','Четвертая Лига','Дивизион А','Дивизион Б']
osnovaniyaKDK=['протокола Матча','видеозаписи Матча','рапорта комиссара','рапорта делегата','рапорта инспектора','рапорта Судьи Матча','письменного заявления членов Президиума МФФ','письменного заявления Клуба', 'письменного заявления Игрока', 'письменного заявления Официального лица','протеста','заявления (рапорта) уполномоченного МФФ лица']
NarushenReglament=['п. 6.1.1.Не уведомлен территориальный орган внутренних дел о проведении Матчей','п. 6.1.2.Не обеспечен общественный порядок и общественную безопасность зрителей и участников','п. 6.1.3.Не назначено лицо ответственное за меры безопасности и общественный порядок, и обеспечить его присутствие в течение всего Матча в пределах Игровой зоны','п. 6.1.4.Не обеспечено наличие служб, отвечающих за безопасность и охрану общественного порядка','п. 6.1.7.Не обеспечено объявление составов Команд, Судей, замен, авторов забитых голов','п. 6.1.10.Не обеспечено наличие системы аудио оповещения на Стадионе','п. 6.1.13.Не обеспечено наличие документации, подтверждающей приемку Стадиона','п. 6.1.16.Не обеспечено наличие Технических зон, а также место для разминки Команд вне Игровой зоны для Команд иной возрастной категории проводящей свой Матч после проводимого Матча','п. 6.1.17.Не предоставлены отдельные раздевалки каждой Команде соперников','п. 6.1.18.Не обеспечено сопровождение Судей, Инспектора от судейской комнаты до футбольного поля и обратно службой охраны общественного порядка','п. 6.1.19.	Не предоставлены Судьям судейская комната и  не обепечено отсутствие посторонних лиц в судейской комнате до, во время, и после Матча','п. 6.1.20. Не предоставлено оборудованное рабочее место в судейской комнате','п. 6.1.21.Не предоставлено: горячий чай и прохладительные напитки для Судей и Инспектора; вода в количестве 1 (один) литр на человека в индивидуальных упаковках (бутилированная)']
NarushKodex=['Статья 70. Неявка на Матч','Статья 71. Задержка начала Матча','Статья 72. Отказ выйти на футбольное поле или продолжить Матч.Отказ покинуть поле','Статья 73. Появление на футбольном поле посторонних лиц','Статья 74. Выход за пределы технической зоны','Статья 75. Курение в пределах Стадиона','Статья 76. Нахождение посторонних лиц в технической зоне и Игровой зоне','Статья 77. Вмешательство в ход Матча','Статья 78. Неспортивное поведение','Статья 78.1. Отказ обмена рукопожатиями','Статья 79. Нарушения, совершенные против Футболистов и Официальных лиц,до, во время и после Матча','Статья 80. Драка','Статья 81. Неправомерное участие в Матче','Статья 82. Провоцирование Зрителей','Статья 83. Необеспечение общественного порядка и безопасности','Статья 84. Прекращение матча','Статья 85. Нанесение материального ущерба командой в предоставленных ей помещениях','Статья 86. Неявка Футболиста или Официального лица в судейскую комнату','Статья 87. Нарушение сроков предоставления информации о времени начала Матча','Статья 88. Неучастие в организационном совещании','Статья 89. Нарушения порядка оформления протокола Матча','Статья 90. Необеспечение условий проведения Матча','Статья 91. Нарушение требований к видеозаписи Матча','Статья 92. Проведение Матча на Стадионе, не имеющем действующего сертификата соответствия МФФ или не имеющем сертификата соответствия МФФ определенной регламентом Соревнований категории','Статья 93. Нарушения требований по медицинскому обеспечению Матча','Статья 94. Нахождение на матче без аккредитации','Статья 95. Невыполнение требований по воспроизведению аудиороликов, аудиорекламы, музыкальному сопровождению матча, по обеспечению работы диктора','Статья 96. Отсутствие баннеров, предоставляемых МФФ','Статья 97. Отсутствие информационного табло','Статья 98. Нахождение лиц, производящих фотосъемку и осуществляющих медицинское обеспечение без специальных накидок (манишек)','Статья 99. Иные нарушения регламента Соревнований','Статья 100. Скандирование зрителями ненормативной лексики и оскорбительных выражений','Статья 101. Бросание зрителями предметов на трибуны, поле и прилегающую к полю территорию','Статья 102. Использование и бросание зрителями пиротехнических изделий','Статья 103. Неправомерные действия зрителей','Статья 104. Подделка документов и использование поддельных документов','Статья 105. Использование Клубом поддельных документов','Статья 106. Нарушения условий участия в Соревновании','Статья 107. Отказ от участия в организационных и торжественных мероприятиях','Статья 108. Нарушение порядка участия и организации спортивных соревнований','Статья 109. Неявка в расположение сборной команды города Москвы']
cvet='#011128'
Style=("Arial", 9)
#def search(word):     # Поиск слова
#ФУНКЦИИ РЕГЛАМЕНТ
def ReglamentKlubnayaLiga():
    z='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\1. Регламенты\\2. Первенство среди спортивных школ\\Регламенты 2020\\ДЮСШ\\Летнее Первенство\\Клубная лига\\РЕГЛАМЕНТ Клубная Лига.docx'
    z1='C:\\Program Files (x86)\\Microsoft Office\\Office14\\WINWORD.EXE'
    subprocess.Popen([z1,z])
def ReglamentPervLiga():
    z='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\1. Регламенты\\2. Первенство среди спортивных школ\\Регламенты 2020\\ДЮСШ\\Летнее Первенство\\Первая лига\\Первая лига.docx'
    z1='C:\\Program Files (x86)\\Microsoft Office\\Office14\\WINWORD.EXE'
    subprocess.Popen([z1,z])
def ReglamentVtorLiga():
    z='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\1. Регламенты\\2. Первенство среди спортивных школ\\Регламенты 2020\\ДЮСШ\\Летнее Первенство\\вторая лига\\Вторая Лига.docx'
    z1='C:\\Program Files (x86)\\Microsoft Office\\Office14\\WINWORD.EXE'
    subprocess.Popen([z1,z])
def ReglamentTretyaLiga():
    z='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\1. Регламенты\\2. Первенство среди спортивных школ\\Регламенты 2020\\ДЮСШ\\Летнее Первенство\\третья лига\\Третья Лига.docx'
    z1='C:\\Program Files (x86)\\Microsoft Office\\Office14\\WINWORD.EXE'
    subprocess.Popen([z1,z])
def ReglamentChetvertayaLiga():
    z='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\1. Регламенты\\2. Первенство среди спортивных школ\\Регламенты 2020\\ДЮСШ\\Летнее Первенство\\четвертая лига\\Четвертая Лига.docx'
    z1='C:\\Program Files (x86)\\Microsoft Office\\Office14\\WINWORD.EXE'
    subprocess.Popen([z1,z])
def ReglamentLFK():
    z='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\1. Регламенты\\3. Чемпионат среди ЛФК\\2020\\Регламент Чемпионата ЛФК 2020.docx'
    z1='C:\\Program Files (x86)\\Microsoft Office\\Office14\\WINWORD.EXE'
    subprocess.Popen([z1,z])
#СОЗДАНИЕ ССЫЛОК И ФАЙЛОВ
def creatfileteam():
    creatteam(creatTEAM.get())
    otvet= Label(window,text='Создан файл {},в папке "10.Организации- участники соревнований" '.format(creatTEAM.get()), font=("Arial", 9))
    otvet.place(x=0, y=520)
def addlinksforschool():
    addlinks(addlinkschool1.get(),addlinkschool2.get())
    otvet1= Label(window,text='Ссылки добавлены'.format(font=Style))
    otvet1.place(x=0, y=510)
def addlinksforLFK():
    addlinksLFK(addlinkschool1.get(),addlinkschool2.get())
    otvet1= Label(window,text='Ссылки добавлены'.format(font=Style))
    otvet1.place(x=0, y=510)
# ФУНКЦИИ
def infoteam():
    z='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\10.Организации- участники соревнований\\'+Komandi[pyperclip.paste()]
    z1='C:\\Program Files (x86)\\Microsoft Office\\Office14\\EXCEL.EXE'
    subprocess.Popen([z1,z])
def creatteam(team):
    os.chdir(put)
    TEAM=openpyxl.Workbook()
    firstsheet=TEAM.active
    firstsheet.title='Правовая информация'
    TEAM.create_sheet(title ='Спортивная информация',index=1)
    TEAM.create_sheet(title ='Аттестация 2020 года',index=2)
    TEAM.create_sheet(title ='Информация о сотрудниках клуба Клуба',index=3)
    TEAM.create_sheet(title ='КДК.КБЭ',index=4)
    pravo=TEAM['Правовая информация']
    sport=TEAM['Спортивная информация']
    attest=TEAM['Аттестация 2020 года']
    works=TEAM['Информация о сотрудниках клуба Клуба']
    sud=TEAM['КДК.КБЭ']
    font = Font(name='Arial', size=8)
    pravospisok=['Наименоваяние организации','Сокращенное наименование','организационно правовая форма','Юридический Адрес','Фактический адрес','Телефон','Официальный сайт','Эл. Почта','ИНН','КПП','ОКПО','ОГРН','Наименование банка','Расчетный счет','Корреспонденсткий счет','БИК','Учредитель','Единоличный Исполнительный Орган','Главный Орган правления','Состав органа правления']
    for i in range(len(pravospisok)):
        pravo['A%s'%(i+1)].font=font
        pravo['A%s'%(i+1)]=pravospisok[i]
    sportspisok=['Название команды','В каких соревнования принимает участие','Стадион' ,'Адрес стадиона']
    for i in range(len(sportspisok)):
        sport['A%s'%(i+1)].font=font
        sport['A%s'%(i+1)]=sportspisok[i]
    attestspisok=['Заявление о проведении аттестации','Заявление о страховании футболистов','Информация о школе','Заявление Спортивной школы о борьбе с расизмом',
                  'Согласие родительского комитета Спортивной школы (в случае наличие родительского комитета)',
                  'Документ, подтверждающий внесение спортивного сооружения во Всероссийский реестр объектов спорта ',
                  'Документ, подтверждающий ввод спортивного сооружения в эксплуатацию',
                  'Документ, подтверждающий право владения/пользования спортивным сооружением для проведения матчей Соревнования',
                  'Инструкция по обеспечению общественного порядка и общественной безопасности на объекте спорта',
                  'Акт обследования и категорирования объекта спорта','План спортивного сооружения',
                  'Информация о сотрудниках Спортивной школы','Заявление о подтверждении соответствия',
                  'Копия Устава, включая все изменения и дополнения, или актуальную редакцию Устава',
                  'Копия документа, подтверждающего государственную регистрацию  Спортивной школы в качестве юридического лица',
                  'Копия документа, подтверждающего постановку Спортивной школы на учет в налоговом органе',
                  'Копия решения, протокола и.тп. уполномоченного органа о создании  Спортивной школы',
                  'Копия решения уполномоченного органа о назначении (избрании на должность) Директора',
                  'Выписка из ЕГРЮЛ','Копия договора аренды помещения, являющегося местом нахождения постоянно действующего исполнительного органа',
                  'Копия договора аренды помещения, являющегося фактическим местом нахождения юридического лица (если фактическое место нахождения не совпадает с юридическим адресом)',
                  'Гарантийное письмо','Документ, подтверждающий оплату платежей, установленных в регламенте Соревнования','Справка Главного бухгалтера РОО МФФ',
                  'Паспорт безопасности']
    for i in range(len(attestspisok)):
        attest['A%s'%(i+1)].font=font
        attest['A%s'%(i+1)]=attestspisok[i]
    workspisok=['Бухгалтер','Тренеры','Cотрудник ответственный за проведение аттестации','Медицинский работник','Другие']
    for i in range(len(workspisok)):
        works['A%s'%(i+1)].font=font
        works['A%s'%(i+1)]=workspisok[i]
    sudspisok=['Орган','Дата и № дела','Сторона в деле','Решение','Оплата штрафа(если есть)']
    for i in range(len(sudspisok)):
        sud['A%s'%(i+1)].font=font
        sud['A%s'%(i+1)]=sudspisok[i]        
    TEAM.save(team+'.'+'xlsx')
    print('готово')
def addlinks(file,papka):
    put='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\4. Аттестация и лицензирование\\Спорт Школы\\2020 год\\Спорт школы\\'+papka
    put1='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\10.Организации- участники соревнований'
    os.chdir(put1)
    spisokpapok=['1. Заявление','4.Заявление о страх', '5. Заявление о расизме', '7. Инструкция', '8.Информа о школе', '9. Инф о сотрудниках', '11. Гарантийное письмо',
                 '12. Оплата взноса', '14. Справка бухгатлера', '15. Внесение  реестр', '16. ВВод в эксплуатац', '17. Право влад стад',
                 '18. Акт обслед и катег', '19. сертиф ворот', '20. План стадиона', '21. Паспорт', '23. Устав', '24. Гос регистр школы',
                 '25. Решение о создании школы', '26. Назначение директора', '27. Егрюл', '28. Регитстрация в налоговой', '29. Офисное помещение']
    team=openpyxl.load_workbook(file+'.'+'xlsx')
    attest=team['Аттестация 2020 года']
    for papk,podpapka,files in os.walk(put):
        try:
            if spisokpapok[0] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=1, column=2).value='file:///'+put+'\\'+os.path.basename(papk)#Заявление о прохождении
                else:
                    attest.cell(row=1, column=2).value='НЕТ'
            if spisokpapok[1] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=2, column=2).value='file:///'+put+'\\'+os.path.basename(papk)#Заявление о страховке
                else:
                    attest.cell(row=2, column=2).value='НЕТ'
            if spisokpapok[2] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=4, column=2).value='file:///'+put+'\\'+os.path.basename(papk)#Заявление о расизме
                else:
                    attest.cell(row=4, column=2).value='НЕТ'
            if spisokpapok[3] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=9, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Инструция
                else:
                    attest.cell(row=9, column=2).value='НЕТ'   
            if spisokpapok[4] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=3, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Информация о школе
                else:
                    attest.cell(row=3, column=2).value='НЕТ'   
            if spisokpapok[5] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=12, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Информация о сотрудниках
                else:
                    attest.cell(row=12, column=2).value='НЕТ'
            if spisokpapok[6] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=22, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Гарантийное письмо
                else:
                    attest.cell(row=22, column=2).value='НЕТ'   
            if spisokpapok[7] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=23, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Оплата взноса
                else:
                    attest.cell(row=23, column=2).value='НЕТ'      
            if spisokpapok[8] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=24, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Справка бухгалтера
                else:
                    attest.cell(row=24, column=2).value='НЕТ'
            if spisokpapok[9] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=6, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Внесение в реестр
                else:
                    attest.cell(row=6, column=2).value='НЕТ'
            if spisokpapok[10] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=7, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Ввод в эксплуатац
                else:
                    attest.cell(row=7, column=2).value='НЕТ'
            if spisokpapok[11] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=8, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Право влад стадионом
                else:
                    attest.cell(row=8, column=2).value='НЕТ'
            if spisokpapok[12] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=10, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Акт обслед и категор
                else:
                    attest.cell(row=10, column=2).value='НЕТ'
            if spisokpapok[14] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=11, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# План Стадиона
                else:
                    attest.cell(row=11, column=2).value='НЕТ'
            if spisokpapok[15] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=25, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Паспорт безопасности
                else:
                    attest.cell(row=25, column=2).value='НЕТ'
            if spisokpapok[16] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=14, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Устав
                else:
                    attest.cell(row=14, column=2).value='НЕТ'
            if spisokpapok[17] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=15, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Гос регистр 
                else:
                    attest.cell(row=15, column=2).value='НЕТ'
            if spisokpapok[18] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=17, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# решение о создании
                else:
                    attest.cell(row=17, column=2).value='НЕТ'
            if spisokpapok[19] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=18, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Назначение директора
                else:
                    attest.cell(row=18, column=2).value='НЕТ'
            if spisokpapok[20] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=19, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Егрюл
                else:
                    attest.cell(row=19, column=2).value='НЕТ'
            if spisokpapok[21] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=16, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Рег в налоговой
                else:
                    attest.cell(row=16, column=2).value='НЕТ'
            if spisokpapok[22] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=20, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# офисное помещение
                else:
                    attest.cell(row=20, column=2).value='НЕТ'
        except FileNotFoundError:
            print('не нашел файл')
    team.save(file+'.'+'xlsx')
def addlinksLFK(file,papka):
    put='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\4. Аттестация и лицензирование\\ЛФК\\2020 год\\'+papka
    put1='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\10.Организации- участники соревнований'
    os.chdir(put1)
    spisokpapok=['1. Заявление','4.Заявление о страх', '5. Заявление о расизме', '7. Инструкция', '8.Информа о школе', '9. Инф о сотрудниках', '11. Гарантийное письмо',
                 '12. Оплата взноса', '14. Справка бухгатлера', '15. Внесение  реестр', '16. ВВод в эксплуатац', '17. Право влад стад',
                 '18. Акт обслед и катег', '19. сертиф ворот', '20. План стадиона', '21. Паспорт', '23. Устав', '24. Гос регистр школы',
                 '25. Решение о создании школы', '26. Назначение директора', '27. Егрюл', '28. Регитстрация в налоговой', '29. Офисное помещение']
    team=openpyxl.load_workbook(file+'.'+'xlsx')
    attest=team['Аттестация 2020 года']
    for papk,podpapka,files in os.walk(put):
        try:
            if spisokpapok[0] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=1, column=2).value='file:///'+put+'\\'+os.path.basename(papk)#Заявление о прохождении
                else:
                    attest.cell(row=1, column=2).value='НЕТ'
            if spisokpapok[1] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=2, column=2).value='file:///'+put+'\\'+os.path.basename(papk)#Заявление о страховке
                else:
                    attest.cell(row=2, column=2).value='НЕТ'
            if spisokpapok[2] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=4, column=2).value='file:///'+put+'\\'+os.path.basename(papk)#Заявление о расизме
                else:
                    attest.cell(row=4, column=2).value='НЕТ'
            if spisokpapok[3] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=9, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Инструция
                else:
                    attest.cell(row=9, column=2).value='НЕТ'   
            if spisokpapok[4] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=3, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Информация о школе
                else:
                    attest.cell(row=3, column=2).value='НЕТ'   
            if spisokpapok[5] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=12, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Информация о сотрудниках
                else:
                    attest.cell(row=12, column=2).value='НЕТ'
            if spisokpapok[6] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=22, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Гарантийное письмо
                else:
                    attest.cell(row=22, column=2).value='НЕТ'   
            if spisokpapok[7] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=23, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Оплата взноса
                else:
                    attest.cell(row=23, column=2).value='НЕТ'      
            if spisokpapok[8] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=24, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Справка бухгалтера
                else:
                    attest.cell(row=24, column=2).value='НЕТ'
            if spisokpapok[9] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=6, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Внесение в реестр
                else:
                    attest.cell(row=6, column=2).value='НЕТ'
            if spisokpapok[10] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=7, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Ввод в эксплуатац
                else:
                    attest.cell(row=7, column=2).value='НЕТ'
            if spisokpapok[11] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=8, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Право влад стадионом
                else:
                    attest.cell(row=8, column=2).value='НЕТ'
            if spisokpapok[12] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=10, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Акт обслед и категор
                else:
                    attest.cell(row=10, column=2).value='НЕТ'
            if spisokpapok[14] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=11, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# План Стадиона
                else:
                    attest.cell(row=11, column=2).value='НЕТ'
            if spisokpapok[15] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=25, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Паспорт безопасности
                else:
                    attest.cell(row=25, column=2).value='НЕТ'
            if spisokpapok[16] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=14, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Устав
                else:
                    attest.cell(row=14, column=2).value='НЕТ'
            if spisokpapok[17] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=15, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Гос регистр 
                else:
                    attest.cell(row=15, column=2).value='НЕТ'
            if spisokpapok[18] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=17, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# решение о создании
                else:
                    attest.cell(row=17, column=2).value='НЕТ'
            if spisokpapok[19] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=18, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Назначение директора
                else:
                    attest.cell(row=18, column=2).value='НЕТ'
            if spisokpapok[20] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=19, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Егрюл
                else:
                    attest.cell(row=19, column=2).value='НЕТ'
            if spisokpapok[21] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=16, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# Рег в налоговой
                else:
                    attest.cell(row=16, column=2).value='НЕТ'
            if spisokpapok[22] in papk:
                if len(os.listdir(put+'\\'+os.path.basename(papk)))>0:
                    attest.cell(row=20, column=2).value='file:///'+put+'\\'+os.path.basename(papk)# офисное помещение
                else:
                    attest.cell(row=20, column=2).value='НЕТ'
        except FileNotFoundError:
            print('не нашел файл')
    team.save(file+'.'+'xlsx')
#ФУНКЦИИ КДК
KnopkiOsnovaniya=[]
KnopkiDlyaNarushKodexa=[]
KnopkiDlyaNarushReglamenta=[]
def KDK():
    #окно
    windowKDK=Toplevel()
    windowKDK.title('КДК')
    windowKDK.geometry('800x800')
    #списки команд
    Ligi = tkinter.StringVar(windowKDK)
    Ligi.set(SpisokLeague[0])
    ViborLigo=tkinter.OptionMenu(windowKDK, Ligi, *SpisokLeague)
    ViborLigo.place(x=20, y=20)
    def dalee():
        if  Ligi.get()=='Клубная Лига':
            variable = tkinter.StringVar(windowKDK)
            variable.set(KlubnayaLigaSpisok[0])
            komanda1 = tkinter.OptionMenu(windowKDK, variable, *KlubnayaLigaSpisok)
            komanda1.config(width=15, font=Style)
            komanda1.place(x=150, y=20)
            variable2 = tkinter.StringVar(windowKDK)
            variable2.set(KlubnayaLigaSpisok[0])
            komanda2= tkinter.OptionMenu(windowKDK, variable2, *KlubnayaLigaSpisok)
            komanda2.config(width=15, font=Style)
            komanda2.place(x=300, y=20)
            def checkStoron():
                Sotrudniki1=[]
                Sotrudniki2=[]
                put1='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\10.Организации- участники соревнований'
                os.chdir(put1)
                team1=openpyxl.load_workbook(Komandi[variable.get()])
                sotrud = team1['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    chelovechki=(str(sotrud['B%s'%(i)].value)).split('\n')
                    for i in chelovechki:
                        print(i)
                        Sotrudniki1.append(i)
                team2=openpyxl.load_workbook(Komandi[variable2.get()])
                sotrud2 = team2['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    chelovechki2=(str(sotrud2['B%s'%(i)].value)).split('\n')
                    for i in chelovechki2:
                        print(i)
                        Sotrudniki2.append(i)
                SotrudKomandi1 = tkinter.StringVar(windowKDK)
                SotrudKomandi1.set(Sotrudniki1[0])
                SotrudKomand1 = tkinter.OptionMenu(windowKDK,  SotrudKomandi1, *Sotrudniki1)
                SotrudKomand1.config(width=30, font=Style)
                SotrudKomand1.place(x=200,y=550)
                SotrudKomandi2 = tkinter.StringVar(windowKDK)
                SotrudKomandi2.set(Sotrudniki2[0])
                SotrudKomand2 = tkinter.OptionMenu(windowKDK,  SotrudKomandi2, *Sotrudniki2)
                SotrudKomand2.config(width=30, font=Style)
                SotrudKomand2.place(x=200,y=600)
                Vizvat['bg']='green'
                def NeVizv():
                    SotrudKomand1.destroy()
                    SotrudKomand2.destroy()
                    NeVizvat['bg']='green'
                    Vizvat['bg']='red'
                NeVizvat=Button(windowKDK,text='Не Вызывать строны',command=NeVizv, font=Style,bg='red')
                NeVizvat.place(x=20, y=600)
            Vizvat= Button(windowKDK,text='Вызвать строны',command=checkStoron, font=Style,bg='red')
            Vizvat.place(x=20, y=550)
        elif Ligi.get()=='Первая Лига':
            variable = tkinter.StringVar(windowKDK)
            variable.set(PervayaLigaLigaSpisok[0])
            komanda1 = tkinter.OptionMenu(windowKDK, variable, *PervayaLigaLigaSpisok)
            komanda1.config(width=15, font=Style)
            komanda1.place(x=100, y=20)
            variable2 = tkinter.StringVar(windowKDK)
            variable2.set(PervayaLigaLigaSpisok[0])
            komanda2= tkinter.OptionMenu(windowKDK, variable2, *PervayaLigaLigaSpisok)
            komanda2.config(width=15, font=Style)
            komanda2.place(x=200, y=20)
            def checkStoron():
                Sotrudniki1=[]
                Sotrudniki2=[]
                put1='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\10.Организации- участники соревнований'
                os.chdir(put1)
                team1=openpyxl.load_workbook(Komandi[variable.get()])
                sotrud = team1['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    Sotrudniki1.append(sotrud['B%s'%(i)].value)
                team2=openpyxl.load_workbook(Komandi[variable2.get()])
                sotrud2 = team2['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    Sotrudniki2.append(sotrud2['B%s'%(i)].value)
                SotrudKomandi1 = tkinter.StringVar(windowKDK)
                SotrudKomandi1.set(Sotrudniki1[0])
                SotrudKomand1 = tkinter.OptionMenu(windowKDK,  SotrudKomandi1, *Sotrudniki1)
                SotrudKomand1.config(width=30, font=Style)
                SotrudKomand1.place(x=200,y=550)
                SotrudKomandi2 = tkinter.StringVar(windowKDK)
                SotrudKomandi2.set(Sotrudniki2[0])
                SotrudKomand2 = tkinter.OptionMenu(windowKDK,  SotrudKomandi2, *Sotrudniki2)
                SotrudKomand2.config(width=30, font=Style)
                SotrudKomand2.place(x=200,y=600)
                Vizvat['bg']='green'
                def NeVizv():
                    SotrudKomand1.destroy()
                    SotrudKomand2.destroy()
                    NeVizvat['bg']='green'
                    Vizvat['bg']='red'
                NeVizvat=Button(windowKDK,text='Не Вызывать строны',command=NeVizv, font=Style,bg='red')
                NeVizvat.place(x=20, y=600)
            Vizvat= Button(windowKDK,text='Вызвать строны',command=checkStoron, font=Style,bg='red')
            Vizvat.place(x=20, y=550)
        elif Ligi.get()=='Вторая Лига':
            variable = tkinter.StringVar(windowKDK)
            variable.set(VtorayaLigaLigaSpisok[0])
            komanda1 = tkinter.OptionMenu(windowKDK, variable, *VtorayaLigaLigaSpisok)
            komanda1.config(width=15, font=Style)
            komanda1.place(x=100, y=20)
            variable2 = tkinter.StringVar(windowKDK)
            variable2.set(VtorayaLigaLigaSpisok[0])
            komanda2= tkinter.OptionMenu(windowKDK, variable2, *VtorayaLigaLigaSpisok)
            komanda2.config(width=15, font=Style)
            komanda2.place(x=200, y=20)
            def checkStoron():
                Sotrudniki1=[]
                Sotrudniki2=[]
                put1='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\10.Организации- участники соревнований'
                os.chdir(put1)
                team1=openpyxl.load_workbook(Komandi[variable.get()])
                sotrud = team1['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    Sotrudniki1.append(sotrud['B%s'%(i)].value)
                team2=openpyxl.load_workbook(Komandi[variable2.get()])
                sotrud2 = team2['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    Sotrudniki2.append(sotrud2['B%s'%(i)].value)
                SotrudKomandi1 = tkinter.StringVar(windowKDK)
                SotrudKomandi1.set(Sotrudniki1[0])
                SotrudKomand1 = tkinter.OptionMenu(windowKDK,  SotrudKomandi1, *Sotrudniki1)
                SotrudKomand1.config(width=30, font=Style)
                SotrudKomand1.place(x=200,y=550)
                SotrudKomandi2 = tkinter.StringVar(windowKDK)
                SotrudKomandi2.set(Sotrudniki2[0])
                SotrudKomand2 = tkinter.OptionMenu(windowKDK,  SotrudKomandi2, *Sotrudniki2)
                SotrudKomand2.config(width=30, font=Style)
                SotrudKomand2.place(x=200,y=600)
                Vizvat['bg']='green'
                def NeVizv():
                    SotrudKomand1.destroy()
                    SotrudKomand2.destroy()
                    NeVizvat['bg']='green'
                    Vizvat['bg']='red'
                NeVizvat=Button(windowKDK,text='Не Вызывать строны',command=NeVizv, font=Style,bg='red')
                NeVizvat.place(x=20, y=600)
            Vizvat= Button(windowKDK,text='Вызвать строны',command=checkStoron, font=Style,bg='red')
            Vizvat.place(x=20, y=550)
        elif Ligi.get()=='Третья Лига':
            variable = tkinter.StringVar(windowKDK)
            variable.set(TretyaLigaLigaSpisok[0])
            komanda1 = tkinter.OptionMenu(windowKDK, variable, *TretyaLigaLigaSpisok)
            komanda1.config(width=15, font=Style)
            komanda1.place(x=100, y=20)
            variable2 = tkinter.StringVar(windowKDK)
            variable2.set(TretyaLigaLigaSpisok[0])
            komanda2= tkinter.OptionMenu(windowKDK, variable2, *TretyaLigaLigaSpisok)
            komanda2.config(width=15, font=Style)
            komanda2.place(x=200, y=20)
            def checkStoron():
                Sotrudniki1=[]
                Sotrudniki2=[]
                put1='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\10.Организации- участники соревнований'
                os.chdir(put1)
                team1=openpyxl.load_workbook(Komandi[variable.get()])
                sotrud = team1['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    Sotrudniki1.append(sotrud['B%s'%(i)].value)
                team2=openpyxl.load_workbook(Komandi[variable2.get()])
                sotrud2 = team2['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    Sotrudniki2.append(sotrud2['B%s'%(i)].value)
                SotrudKomandi1 = tkinter.StringVar(windowKDK)
                SotrudKomandi1.set(Sotrudniki1[0])
                SotrudKomand1 = tkinter.OptionMenu(windowKDK,  SotrudKomandi1, *Sotrudniki1)
                SotrudKomand1.config(width=30, font=Style)
                SotrudKomand1.place(x=200,y=550)
                SotrudKomandi2 = tkinter.StringVar(windowKDK)
                SotrudKomandi2.set(Sotrudniki2[0])
                SotrudKomand2 = tkinter.OptionMenu(windowKDK,  SotrudKomandi2, *Sotrudniki2)
                SotrudKomand2.config(width=30, font=Style)
                SotrudKomand2.place(x=200,y=600)
                Vizvat['bg']='green'
                def NeVizv():
                    SotrudKomand1.destroy()
                    SotrudKomand2.destroy()
                    NeVizvat['bg']='green'
                    Vizvat['bg']='red'
                NeVizvat=Button(windowKDK,text='Не Вызывать строны',command=NeVizv, font=Style,bg='red')
                NeVizvat.place(x=20, y=600)
            Vizvat= Button(windowKDK,text='Вызвать строны',command=checkStoron, font=Style,bg='red')
            Vizvat.place(x=20, y=550)
        elif Ligi.get()=='Четвертая Лига':
            variable = tkinter.StringVar(windowKDK)
            variable.set(ChetvetrtayaLigaLigaSpisok[0])
            komanda1 = tkinter.OptionMenu(windowKDK, variable, *ChetvetrtayaLigaLigaSpisok)
            komanda1.config(width=15, font=Style)
            komanda1.place(x=100, y=20)
            variable2 = tkinter.StringVar(windowKDK)
            variable2.set(ChetvetrtayaLigaLigaSpisok[0])
            komanda2= tkinter.OptionMenu(windowKDK, variable2, *ChetvetrtayaLigaLigaSpisok)
            komanda2.config(width=15, font=Style)
            komanda2.place(x=200, y=20)
            def checkStoron():
                Sotrudniki1=[]
                Sotrudniki2=[]
                put1='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\10.Организации- участники соревнований'
                os.chdir(put1)
                team1=openpyxl.load_workbook(Komandi[variable.get()])
                sotrud = team1['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    Sotrudniki1.append(sotrud['B%s'%(i)].value)
                team2=openpyxl.load_workbook(Komandi[variable2.get()])
                sotrud2 = team2['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    Sotrudniki2.append(sotrud2['B%s'%(i)].value)
                SotrudKomandi1 = tkinter.StringVar(windowKDK)
                SotrudKomandi1.set(Sotrudniki1[0])
                SotrudKomand1 = tkinter.OptionMenu(windowKDK,  SotrudKomandi1, *Sotrudniki1)
                SotrudKomand1.config(width=30, font=Style)
                SotrudKomand1.place(x=200,y=550)
                SotrudKomandi2 = tkinter.StringVar(windowKDK)
                SotrudKomandi2.set(Sotrudniki2[0])
                SotrudKomand2 = tkinter.OptionMenu(windowKDK,  SotrudKomandi2, *Sotrudniki2)
                SotrudKomand2.config(width=30, font=Style)
                SotrudKomand2.place(x=200,y=600)
                Vizvat['bg']='green'
                def NeVizv():
                    SotrudKomand1.destroy()
                    SotrudKomand2.destroy()
                    NeVizvat['bg']='green'
                    Vizvat['bg']='red'
                NeVizvat=Button(windowKDK,text='Не Вызывать строны',command=NeVizv, font=Style,bg='red')
                NeVizvat.place(x=20, y=600)
            Vizvat= Button(windowKDK,text='Вызвать строны',command=checkStoron, font=Style,bg='red')
            Vizvat.place(x=20, y=550)
        elif Ligi.get()=='Дивизион А':
            variable = tkinter.StringVar(windowKDK)
            variable.set(DivizASpisok[0])
            komanda1 = tkinter.OptionMenu(windowKDK, variable, *DivizASpisok)
            komanda1.config(width=15, font=Style)
            komanda1.place(x=100, y=20)
            variable2 = tkinter.StringVar(windowKDK)
            variable2.set(DivizASpisok[0])
            komanda2= tkinter.OptionMenu(windowKDK, variable2, *DivizASpisok)
            komanda2.config(width=15, font=Style)
            komanda2.place(x=200, y=20)
            def checkStoron():
                Sotrudniki1=[]
                Sotrudniki2=[]
                put1='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\10.Организации- участники соревнований'
                os.chdir(put1)
                team1=openpyxl.load_workbook(Komandi[variable.get()])
                sotrud = team1['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    Sotrudniki1.append(sotrud['B%s'%(i)].value)
                team2=openpyxl.load_workbook(Komandi[variable2.get()])
                sotrud2 = team2['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    Sotrudniki2.append(sotrud2['B%s'%(i)].value)
                SotrudKomandi1 = tkinter.StringVar(windowKDK)
                SotrudKomandi1.set(Sotrudniki1[0])
                SotrudKomand1 = tkinter.OptionMenu(windowKDK,  SotrudKomandi1, *Sotrudniki1)
                SotrudKomand1.config(width=30, font=Style)
                SotrudKomand1.place(x=200,y=550)
                SotrudKomandi2 = tkinter.StringVar(windowKDK)
                SotrudKomandi2.set(Sotrudniki2[0])
                SotrudKomand2 = tkinter.OptionMenu(windowKDK,  SotrudKomandi2, *Sotrudniki2)
                SotrudKomand2.config(width=30, font=Style)
                SotrudKomand2.place(x=200,y=600)
                Vizvat['bg']='green'
                def NeVizv():
                    SotrudKomand1.destroy()
                    SotrudKomand2.destroy()
                    NeVizvat['bg']='green'
                    Vizvat['bg']='red'
                NeVizvat=Button(windowKDK,text='Не Вызывать строны',command=NeVizv, font=Style,bg='red')
                NeVizvat.place(x=20, y=600)
            Vizvat= Button(windowKDK,text='Вызвать строны',command=checkStoron, font=Style,bg='red')
            Vizvat.place(x=20, y=550)
        elif Ligi.get()=='Дивизион Б':
            variable = tkinter.StringVar(windowKDK)
            variable.set(DicizBSpisok[0])
            komanda1 = tkinter.OptionMenu(windowKDK, variable, *DicizBSpisok)
            komanda1.config(width=15, font=Style)
            komanda1.place(x=100, y=20)
            variable2 = tkinter.StringVar(windowKDK)
            variable2.set(DicizBSpisok[0])
            komanda2= tkinter.OptionMenu(windowKDK, variable2, *DicizBSpisok)
            komanda2.config(width=15, font=Style)
            komanda2.place(x=200, y=20)
            def checkStoron():
                Sotrudniki1=[]
                Sotrudniki2=[]
                put1='C:\\Users\\Alexander\\Desktop\\Рабочая папка\\1. МФФ\\2. Документы связанные с осноной деятельностью\\10.Организации- участники соревнований'
                os.chdir(put1)
                team1=openpyxl.load_workbook(Komandi[variable.get()])
                sotrud = team1['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    Sotrudniki1.append(sotrud['B%s'%(i)].value)
                team2=openpyxl.load_workbook(Komandi[variable2.get()])
                sotrud2 = team2['Информация о сотрудниках клуба ']
                for i in range (1,6):
                    Sotrudniki2.append(sotrud2['B%s'%(i)].value)
                SotrudKomandi1 = tkinter.StringVar(windowKDK)
                SotrudKomandi1.set(Sotrudniki1[0])
                SotrudKomand1 = tkinter.OptionMenu(windowKDK,  SotrudKomandi1, *Sotrudniki1)
                SotrudKomand1.config(width=30, font=Style)
                SotrudKomand1.place(x=200,y=550)
                SotrudKomandi2 = tkinter.StringVar(windowKDK)
                SotrudKomandi2.set(Sotrudniki2[0])
                SotrudKomand2 = tkinter.OptionMenu(windowKDK,  SotrudKomandi2, *Sotrudniki2)
                SotrudKomand2.config(width=30, font=Style)
                SotrudKomand2.place(x=200,y=600)
                Vizvat['bg']='green'
                def NeVizv():
                    SotrudKomand1.destroy()
                    SotrudKomand2.destroy()
                    NeVizvat['bg']='green'
                    Vizvat['bg']='red'
                NeVizvat=Button(windowKDK,text='Не Вызывать строны',command=NeVizv, font=Style,bg='red')
                NeVizvat.place(x=20, y=600)
            Vizvat= Button(windowKDK,text='Вызвать строны',command=checkStoron, font=Style,bg='red')
            Vizvat.place(x=20, y=550)
    Dalee=Button(windowKDK,text='Далее',command=dalee,font=Style)
    Dalee.place(x=450,y=20)
    Tour = Entry(windowKDK,width=10)  
    Tour.place(x=45, y=60)
    ToutLabel=Label(windowKDK,text='Тур',font=Style)
    ToutLabel.place(x=20,y=60)
    NaOsnovaniiLabel=Label(windowKDK,text='На основании',font=Style)
    NaOsnovaniiLabel.place(x=20,y=80)
    NaOsnovanii=Text(windowKDK,width=25, height=5, wrap=WORD)
    NaOsnovanii.place(x=300, y=80)
    NaOsnovaniiibor = tkinter.StringVar(windowKDK)
    NaOsnovaniiibor.set(osnovaniyaKDK[0])
    NaOsnovaniiibor1 = tkinter.OptionMenu(windowKDK, NaOsnovaniiibor, *osnovaniyaKDK)
    NaOsnovaniiibor1.config(width=15, font=Style)
    NaOsnovaniiibor1.place(x=150, y=80)
    def AddOsovanie():
        NovKnop='NaOsnovaniiibor'+str(len(KnopkiOsnovaniya))
        KnopkiOsnovaniya.append(NovKnop)
        NovKnop = tkinter.StringVar(windowKDK)
        NovKnop.set(osnovaniyaKDK[0])
        mesto=str(NovKnop)+str(len(KnopkiOsnovaniya))
        mesto = tkinter.OptionMenu(windowKDK, NovKnop, *osnovaniyaKDK)
        mesto.config(width=15, font=Style)
        mesto.place(x=NaOsnovaniiibor1.winfo_x(), y=NaOsnovaniiibor1.winfo_y()+(30*(len (KnopkiOsnovaniya))))
    ADD=Button(windowKDK,text='Добавить',command=AddOsovanie, font=Style)
    ADD.place(x=20,y=110)
    KodexNarush = tkinter.StringVar(windowKDK)
    KodexNarush.set(NarushKodex[0])
    KodexNarush1 = tkinter.OptionMenu(windowKDK, KodexNarush, *NarushKodex)
    KodexNarush1.config(width=50, font=Style)
    KodexNarush1.place(x=20,y=200)
    def AddNarushenie():
        NovNarush='KodexNarush'+str(len(KnopkiDlyaNarushKodexa))
        KnopkiDlyaNarushKodexa.append(NovNarush)
        NovNarush = tkinter.StringVar(windowKDK)
        NovNarush.set(NarushKodex[0])
        maisto=str(NovNarush)+str(len(KnopkiDlyaNarushKodexa))
        maisto = tkinter.OptionMenu(windowKDK, NovNarush, *NarushKodex)
        maisto.config(width=50, font=Style)
        maisto.place(x=KodexNarush1.winfo_x(), y=KodexNarush1.winfo_y()+(30*(len (KnopkiDlyaNarushKodexa))))
    PravReglamenta = tkinter.StringVar(windowKDK)
    PravReglamenta.set(NarushenReglament[0])
    PravReglamenta1 = tkinter.OptionMenu(windowKDK, PravReglamenta, *NarushenReglament)
    PravReglamenta1.config(width=50, font=Style)
    PravReglamenta1.place(x=20,y=350)
    def AddNarushenie1():
        NovNarush1='KodexNarush'+str(len(KnopkiDlyaNarushReglamenta))
        KnopkiDlyaNarushReglamenta.append(NovNarush1)
        NovNarush1 = tkinter.StringVar(windowKDK)
        NovNarush1.set(NarushenReglament[0])
        maisto1=str(NovNarush1)+str(len(KnopkiDlyaNarushReglamenta))
        maisto1 = tkinter.OptionMenu(windowKDK, NovNarush1, *NarushenReglament)
        maisto1.config(width=50, font=Style)
        maisto1.place(x=PravReglamenta1.winfo_x(), y=PravReglamenta1.winfo_y()+(30*(len (KnopkiDlyaNarushReglamenta))))
    ADD2=Button(windowKDK,text='Добавить',command=AddNarushenie, font=Style)
    ADD2.place(x=440,y=200)
    ADD3=Button(windowKDK,text='Добавить',command=AddNarushenie1, font=Style)
    ADD3.place(x=440,y=350)
    Date = Entry(windowKDK,width=10)  
    Date.place(x=120, y=500)
    DateLabel=Label(windowKDK,text='Дата Заседания',font=Style)
    DateLabel.place(x=20,y=500)


   
    windowKDK.mainloop()
# ИНТЕРФЕЙС
# №1 ОКНО
window=Tk()
window.title('База команд МФФ')
window.geometry('1366x768')
image2 =Image.open('oboi.jpg')
image1 = ImageTk.PhotoImage(image2)
bg_label = tkinter.Label(window, image = image1)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)    
# ТАБЛИЦА КОМАНД
# КЛУБНАЯ ЛИГА
KlubnayaLiga = Button(window, text="Клубная лига", font=Style,command=ReglamentKlubnayaLiga) 
KlubnayaLiga.place(x=0, y=0)
#РЕГЛАМЕНТ
PeriodKlubLiga = Label(window, text="08/20-??/21",font=Style) 
PeriodKlubLiga.place(x=0, y=26)
#КОМАНДЫ #Перемещение в зависимости от таблицы
Lokomoti='Локомотив-1'
Lokomotiv = Button(window, font=Style,text="Локомотив-1",command=lambda:[pyperclip.copy(Lokomoti),infoteam()])
Lokomotiv.place(x=0, y=49)
Lokomotiv2 = Button(window,font=Style, text="Локомотив-2",command=lambda:[pyperclip.copy(Lokomoti),infoteam()])
Lokomotiv2.place(x=0, y=74)
Sparta='Спартак'
Spartak = Button(window, font=Style,text="Спартак",command=lambda:[pyperclip.copy(Sparta),infoteam()])
Spartak.place(x=0, y=99)
Dinam='Динамо'
Dinamo=Button(window,font=Style, text="Динамо",command=lambda:[pyperclip.copy(Dinam),infoteam()])
Dinamo.place(x=0, y=124)
CSKa='ЦСКА'
CSKA=Button(window, font=Style,text="ЦСКА",command=lambda:[pyperclip.copy(CSKa),infoteam()])
CSKA.place(x=0, y=149)
Chertanov='Чертаново'
Chertanovo=Button(window,font=Style, text="Чертаново",command=lambda:[pyperclip.copy(Chertanov),infoteam()])
Chertanovo.place(x=0, y=174)
FSH='ФШМ'
FSHM=Button(window, font=Style,text="ФШМ",command=lambda:[pyperclip.copy(FSH),infoteam()])
FSHM.place(x=0, y=199)
Strogin='Строгино'
Strogino=Button(window, font=Style,text="Строгино",command=lambda:[pyperclip.copy(Strogin),infoteam()])
Strogino.place(x=0, y=224)
Himk='Химки'
Himki=Button(window,font=Style,text="Химки",command=lambda:[pyperclip.copy(Himk),infoteam()])
Himki.place(x=0, y=249)
Rodin='Родина'
Rodina=Button(window, font=Style,text="Родина",command=lambda:[pyperclip.copy(Rodin),infoteam()])
Rodina.place(x=0, y=274)
# ПЕРВАЯ ЛИГА
PervayaLiga = Button(window, text="Первая лига", font=Style,command=ReglamentPervLiga) 
PervayaLiga.place(x=110, y=0)
#РЕГЛАМЕНТ
ReglamentPervLiga = Label(window, text="08/20-??/21",font=Style) 
ReglamentPervLiga.place(x=110, y=26)
#КОМАНДЫ #Перемещение в зависимости от таблицы
Spart='Спартак-2'
Spartak2 = Button(window,font=Style, text="Спартак-2",command=lambda:[pyperclip.copy(Spart),infoteam()])
Spartak2.place(x=110, y=49)
TorpedoMosKomSpor='ТорпедоМосКомСпорт'
TorpedoMosKomSport=Button(window,font=Style, text="ТорпедоМосКомСпорт",command=lambda:[pyperclip.copy(TorpedoMosKomSpor),infoteam()])
TorpedoMosKomSport.place(x=110, y=74)
Smen='Смена'
Smena=Button(window, font=Style,text="Смена",command=lambda:[pyperclip.copy(Smen),infoteam()])
Smena.place(x=110, y=99)
Soko='Сокол'
Sokol=Button(window, font=Style,text="Сокол",command=lambda:[pyperclip.copy(Soko),infoteam()])
Sokol.place(x=110, y=124)
TrudRezerv='Трудовые резервы'
TrudovieRezervi=Button(window,font=Style, text="Трудовые резервы",command=lambda:[pyperclip.copy(TrudRezerv),infoteam()])
TrudovieRezervi.place(x=110, y=149)
KrilyaSoveto='Крылья Советов'
KrilyaSovetov=Button(window, font=Style,text="Крылья Советов",command=lambda:[pyperclip.copy(KrilyaSoveto),infoteam()])
KrilyaSovetov.place(x=110, y=174)
Strogin2='Строгино-2'
Strogino2= Button(window,font=Style, text="Строгино-2",command=lambda:[pyperclip.copy(Strogin2),infoteam()])
Strogino2.place(x=110, y=199)
Priali='Приалит'
Prialit= Button(window,font=Style, text="Приалит",command=lambda:[pyperclip.copy(Priali),infoteam()])
Prialit.place(x=110, y=224)
Rosic='Росич'
Rosich= Button(window,font=Style, text="Росич",command=lambda:[pyperclip.copy(Rosic),infoteam()])
Rosich.place(x=110, y=249)
fsh2='ФШМ-2'
fshm2= Button(window,font=Style, text="ФШМ-2",command=lambda:[pyperclip.copy(fsh2),infoteam()])
fshm2.place(x=110, y=274)
Sputni='Спутник'
Sputnik= Button(window,font=Style, text="Спутник",command=lambda:[pyperclip.copy(Sputni),infoteam()])
Sputnik.place(x=110, y=299)
Chertnanov2='Чертаново-2'
Chertnanovo2=Button(window,font=Style, text="Чертаново-2",command=lambda:[pyperclip.copy(Chertnanov2),infoteam()])
Chertnanovo2.place(x=110, y=324)
# ВТОРАЯ ЛИГА
VtorayaLiga = Button(window, text="Вторая лига", font=Style,command=ReglamentVtorLiga) 
VtorayaLiga.place(x=270, y=0)
#РЕГЛАМЕНТ
ReglamentVtorLiga = Label(window, text="08/20-??/21",font=Style) 
ReglamentVtorLiga.place(x=270, y=26)
#КОМАНДЫ #Перемещение в зависимости от таблицы
Burevestni='Буревестник'
Burevestnik=Button(window,font=Style, text="Буревестник",command=lambda:[pyperclip.copy(Burevestni),infoteam()])
Burevestnik.place(x=270, y=49)
Kuncev='Кунцево'
Kuncevo=Button(window, font=Style,text="Кунцево",command=lambda:[pyperclip.copy(Kuncev),infoteam()])
Kuncevo.place(x=270, y=74)
Dinam2='Динамо-2'
Dinamo2=Button(window,font=Style, text="Динамо-2",command=lambda:[pyperclip.copy(Dinam2),infoteam()])
Dinamo2.place(x=270, y=99)
Krasnogvardee='Красногвардеец'
Krasnogvardeec=Button(window, font=Style,text="Красногвардеец",command=lambda:[pyperclip.copy(Krasnogvardee),infoteam()])
Krasnogvardeec.place(x=270, y=124)
Megasfer='Мегасфера'
Megasfera=Button(window,font=Style, text="Мегасфера",command=lambda:[pyperclip.copy(Megasfer),infoteam()])
Megasfera.place(x=270, y=149)
Savelovskay='Савеловская'
Savelovskaya=Button(window,font=Style, text="Савеловская",command=lambda:[pyperclip.copy(Savelovskay),infoteam()])
Savelovskaya.place(x=270, y=174)
Moskvic='Москвич'
Moskvich=Button(window, font=Style,text="Москвич",command=lambda:[pyperclip.copy(Moskvic),infoteam()])
Moskvich.place(x=270, y=199)
sh='СШ-4'
ssh=Button(window,font=Style, text="СШ-4",command=lambda:[pyperclip.copy(sh),infoteam()])
ssh.place(x=270, y=224)
sas='САШ Олимпик'
sash=Button(window, font=Style,text="САШ Олимпик",command=lambda:[pyperclip.copy(sas),infoteam()])
sash.place(x=270, y=249)
Torped2='Торпедо-2'
Torpedo2=Button(window, font=Style,text="Торпедо-2",command=lambda:[pyperclip.copy(Torped2),infoteam()])
Torpedo2.place(x=270, y=274)
smen2='Смена -2'
smena2=Button(window,font=Style, text="Смена -2",command=lambda:[pyperclip.copy(smen2),infoteam()])
smena2.place(x=270, y=299)
molniy='Молния'
molniya=Button(window,font=Style, text="Молния",command=lambda:[pyperclip.copy(molniy),infoteam()])
molniya.place(x=270, y=324)
# ТРЕТЬЯ ЛИГА
TretyaLiga = Button(window, text="Третья лига", font=Style,command=ReglamentTretyaLiga) 
TretyaLiga.place(x=390, y=0)
#РЕГЛАМЕНТ
ReglamentTretyaLiga = Label(window, text="08/20-??/21",font=Style) 
ReglamentTretyaLiga.place(x=390, y=26)
#КОМАНДЫ #Перемещение в зависимости от таблицы
KrilyaSoveto2='Крылья Советов -2'
KrilyaSovetov2=Button(window,font=Style, text="Крылья Советов -2",command=lambda:[pyperclip.copy(KrilyaSoveto2),infoteam()])
KrilyaSovetov2.place(x=390, y=49)
Savelovsay2='Савеловская-2'
Savelovsaya2=Button(window, font=Style,text="Савеловская-2",command=lambda:[pyperclip.copy(Savelovsay2),infoteam()])
Savelovsaya2.place(x=390, y=74)
VorobyovoGor='Воробьевы Горы'
VorobyovoGori=Button(window, font=Style,text="Воробьевы Горы",command=lambda:[pyperclip.copy(VorobyovoGor),infoteam()])
VorobyovoGori.place(x=390, y=99)
KrilyaSoveto3='Крылья Советов-3'
KrilyaSovetov3=Button(window, font=Style,text="Крылья Советов-3",command=lambda:[pyperclip.copy(KrilyaSoveto3),infoteam()])
KrilyaSovetov3.place(x=390, y=124)
TrudovieRezerv2='Трудовые резервы-2'
TrudovieRezervi2=Button(window,font=Style, text="Трудовые резервы-2",command=lambda:[pyperclip.copy(TrudovieRezerv2),infoteam()])
TrudovieRezervi2.place(x=390, y=149)
Tushin='СШ №101 Тушино'
Tushino=Button(window, font=Style,text="СШ №101 Тушино",command=lambda:[pyperclip.copy(Tushin),infoteam()])
Tushino.place(x=390, y=174)
Tcaricin='Царицыно'
Tcaricino=Button(window, font=Style,text="Царицыно",command=lambda:[pyperclip.copy(Tcaricin),infoteam()])
Tcaricino.place(x=390, y=199)
Ssh8='СШ №82'
Ssh82=Button(window,font=Style, text="СШ №82",command=lambda:[pyperclip.copy(Ssh8),infoteam()])
Ssh82.place(x=390, y=224)
Troic='Троицк'
Troick=Button(window, font=Style,text="Троицк",command=lambda:[pyperclip.copy(Troic),infoteam()])
Troick.place(x=390, y=249)
Krasnogvardee2='Красногвардеец-2'
Krasnogvardeec2=Button(window,font=Style, text="Красногвардеец-2",command=lambda:[pyperclip.copy(Krasnogvardee2),infoteam()])
Krasnogvardeec2.place(x=390, y=274)
Mitin='Митино'
Mitino=Button(window, font=Style,text="Митино",command=lambda:[pyperclip.copy(Mitin),infoteam()])
Mitino.place(x=390, y=299)
Vitya='Витязь'
Vityaz=Button(window,font=Style, text="Витязь",command=lambda:[pyperclip.copy(Vitya),infoteam()])
Vityaz.place(x=390, y=324)
SpartaSviblov='Спарта-Свиблово'
SpartaSviblovo=Button(window,font=Style, text="Спарта-Свиблово",command=lambda:[pyperclip.copy(SpartaSviblov),infoteam()])
SpartaSviblovo.place(x=390, y=349)
SmenaUnio='Смена-юниор'
SmenaUnior=Button(window, font=Style,text="Смена-юниор",command=lambda:[pyperclip.copy(SmenaUnio),infoteam()])
SmenaUnior.place(x=390, y=374)
# ЧЕТВЕРТАЯ ЛИГА
ChetvertayaLiga = Button(window, text="Четвертая лига", font=Style,command=ReglamentChetvertayaLiga) 
ChetvertayaLiga.place(x=540, y=0)
#РЕГЛАМЕНТ
ReglamentChetvertayaLiga = Label(window, text="08/20-??/21",font=Style) 
ReglamentChetvertayaLiga.place(x=540, y=26)
#КОМАНДЫ #Перемещение в зависимости от таблицы
Avangar='Авангард'
Avangard=Button(window,font=Style, text="Авангард",command=lambda:[pyperclip.copy(Avangar),infoteam()])
Avangard.place(x=540, y=49)
Od8='ОД-80'
Od80=Button(window, font=Style,text="ОД-80",command=lambda:[pyperclip.copy(Od8),infoteam()])
Od80.place(x=540, y=74)
MV='МВА'
MVA=Button(window,font=Style, text="МВА",command=lambda:[pyperclip.copy(MV),infoteam()])
MVA.place(x=540, y=99)
Unitr='Юнитра'
Unitra=Button(window, font=Style,text="Юнитра",command=lambda:[pyperclip.copy(Unitr),infoteam()])
Unitra.place(x=540, y=124)
Soko2='Сокол-2'
Sokol2=Button(window, font=Style,text="Сокол-2",command=lambda:[pyperclip.copy(Soko2),infoteam()])
Sokol2.place(x=540, y=149)
Megasfer2='Мегасфера-2'
Megasfera2=Button(window,font=Style, text="Мегасфера-2",command=lambda:[pyperclip.copy(Megasfer2),infoteam()])
Megasfera2.place(x=540, y=174)
AkademiaSalov='Лужники'
AkademiaSalova=Button(window,font=Style, text="Лужники",command=lambda:[pyperclip.copy(AkademiaSalov),infoteam()])
AkademiaSalova.place(x=540, y=199)
Kuncev2='Кунцево-2'
Kuncevo2=Button(window, font=Style,text="Кунцево-2",command=lambda:[pyperclip.copy(Kuncev2),infoteam()])
Kuncevo2.place(x=540, y=224)
Torped='ДЮСШ Торпедо'
Torpedo=Button(window,font=Style, text="ДЮСШ Торпедо",command=lambda:[pyperclip.copy(Torped),infoteam()])
Torpedo.place(x=540, y=249)
Buduge='Будущее'
Budugee=Button(window, font=Style,text="Будущее",command=lambda:[pyperclip.copy(Buduge),infoteam()])
Budugee.place(x=540, y=274)
Krepos='Крепость'
Krepost=Button(window, font=Style,text="Крепость",command=lambda:[pyperclip.copy(Krepos),infoteam()])
Krepost.place(x=540, y=299)
Lobny='Лобня'
Lobnya=Button(window, font=Style,text="Лобня",command=lambda:[pyperclip.copy(Lobny),infoteam()])
Lobnya.place(x=540, y=324)
Savelovsay3='Савеловская-3'
Savelovsaya3=Button(window, font=Style,text="Савеловская-3",command=lambda:[pyperclip.copy(Savelovsay3),infoteam()])
Savelovsaya3.place(x=540, y=349)
Savelovsaya3['state'] = 'disabled'
Luc='Луч'
Luch=Button(window, font=Style,text="Луч",command=lambda:[pyperclip.copy(Luc),infoteam()])
Luch.place(x=540, y=374)
Luch['state'] = 'disabled'
Admira='Адмирал'
Admiral=Button(window,font=Style, text="Адмирал",command=lambda:[pyperclip.copy(Admira),infoteam()])
Admiral.place(x=540, y=399)
Admiral['state'] = 'disabled'
FCA='ФК АС'
FCAS=Button(window,font=Style, text="ФК АС",command=lambda:[pyperclip.copy(FCA),infoteam()])
FCAS.place(x=540, y=424)
FCAS['state'] = 'disabled'
# ЛФК ЛИГА
LFK = Button(window, text="Чемпионат среди ЛФК",command=ReglamentLFK,font=("Arial", 8))
LFK.place(x=670, y=0)
LFKDIVA = Label(window, text='Дивизион 3 "А"', font=Style) 
LFKDIVA.place(x=670, y=26)
PeriodLFK = Label(window, text='23/07/20-??/21', font=Style)
PeriodLFK.place(x=794, y=0)
LFKDIVB = Label(window, text='Дивизион 4 "Б"', font=Style) 
LFKDIVB.place(x=790, y=26)
#КОМАНДЫ ДИВИЗИОН А #Перемещение в зависимости от таблицы
RodinLF='Родина-2'
RodinLFK=Button(window,font=Style, text="Родина-2",command=lambda:[pyperclip.copy(RodinLF),infoteam()])
RodinLFK.place(x=670, y=49)
RosicLF='Росич'
RosicLFK=Button(window,font=Style, text="Росич",command=lambda:[pyperclip.copy(RosicLF),infoteam()])
RosicLFK.place(x=670, y=74)
SetunKuncev='Сетунь-Кунцево'
SetunKuncevo=Button(window,font=Style, text="Сетунь-Кунцево",command=lambda:[pyperclip.copy(SetunKuncev),infoteam()])
SetunKuncevo.place(x=670, y=99)
BurevestniLFK='Буревестник'
BurevestnikLFK=Button(window, font=Style,text="Буревестник",command=lambda:[pyperclip.copy(BurevestniLFK),infoteam()])
BurevestnikLFK.place(x=670, y=124)
LetniyDogdi='Летний дождик'
LetniyDogdik=Button(window, font=Style,text="Летний дождик",command=lambda:[pyperclip.copy(LetniyDogdi),infoteam()])
LetniyDogdik.place(x=670, y=149)
Zeni='Зенит'
Zenit=Button(window, font=Style,text="Зенит",command=lambda:[pyperclip.copy(Zeni),infoteam()])
Zenit.place(x=670, y=174)
FSHMLF='ФШМ'
FSHMLFK=Button(window,font=Style, text="ФШМ",command=lambda:[pyperclip.copy(FSHMLF),infoteam()])
FSHMLFK.place(x=670, y=199)
SmenLFK='Смена'
SmenaLFK=Button(window,font=Style, text="Смена",command=lambda:[pyperclip.copy(SmenLFK),infoteam()])
SmenaLFK.place(x=670, y=224)
Geraklio='ГЕРАКЛИОН'
Geraklion=Button(window, font=Style,text="ГЕРАКЛИОН",command=lambda:[pyperclip.copy(Geraklio),infoteam()])
Geraklion.place(x=670, y=249)
SSH7='СШ-75'
SSH75=Button(window, font=Style,text="СШ-75",command=lambda:[pyperclip.copy(SSH7),infoteam()])
SSH75.place(x=670, y=274)
SSH75['state'] = 'disabled'
Zelenogra='Зеленоград'
Zelenograd=Button(window, font=Style,text="Зеленоград",command=lambda:[pyperclip.copy(Zelenogra),infoteam()])
Zelenograd.place(x=670, y=299)
Gefes='Гефест'
Gefest=Button(window, font=Style,text="Гефест",command=lambda:[pyperclip.copy(Gefes),infoteam()])
Gefest.place(x=670, y=324)
TrudRezLF='Трудовые резервы'
TrudRezLFK=Button(window, font=Style,text="Трудовые резервы",command=lambda:[pyperclip.copy(TrudRezLF),infoteam()])
TrudRezLFK.place(x=670, y=349)
LokLFK='ЛОКОМОТИВЛФК'
LokoLFK=Button(window,font=Style, text="ЛОКОМОТИВ",command=lambda:[pyperclip.copy(LokLFK),infoteam()])
LokoLFK.place(x=670, y=374)
StroginM='Строгино-м'
StroginoM=Button(window,font=Style, text="Строгино-м",command=lambda:[pyperclip.copy(StroginM),infoteam()])
StroginoM.place(x=670, y=399)
SportAkadKlu='Спортакадемклуб'
SportAkadKlub=Button(window,font=Style, text="Спортакадемклуб",command=lambda:[pyperclip.copy(SportAkadKlu),infoteam()])
SportAkadKlub.place(x=670, y=424)
SportAkadKlub['state'] = 'disabled'
Sahalin='Сахалинец'
Sahalinec=Button(window,font=Style, text='Сахалинец',command=lambda:[pyperclip.copy(Sahalin),infoteam()])
Sahalinec.place(x=670, y=449)
#КОМАНДЫ ДИВИЗИОН Б #Перемещение в зависимости от таблицы
TroicLFK='Троицк'
TroickLFK=Button(window, font=Style,text="Троицк",command=lambda:[pyperclip.copy(TroicLFK),infoteam()])
TroickLFK.place(x=790, y=49)
ss4LFK='СШ-4 Легион'
ssh4LFK=Button(window,font=Style, text="СШ-4 Легион",command=lambda:[pyperclip.copy(ss4LFK),infoteam()])
ssh4LFK.place(x=790, y=74)
KrilyaAlyan='Крылья Советов-Альянс'
KrilyaAlyans=Button(window, font=Style,text="Крылья Советов-Альянс",command=lambda:[pyperclip.copy(KrilyaAlyan),infoteam()])
KrilyaAlyans.place(x=790, y=99)
KrilyaAlyans['state'] = 'disabled'
RosicLFK='Росич-2'
RosichLFK=Button(window,font=Style, text="Росич-2",command=lambda:[pyperclip.copy(RosicLFK),infoteam()])
RosichLFK.place(x=790, y=124)
Centav='Центавр'
Centavr=Button(window,font=Style, text="Центавр",command=lambda:[pyperclip.copy(Centav),infoteam()])
Centavr.place(x=790, y=149)
TorpedoLF='Торпедо' 
TorpedoLFK=Button(window,font=Style, text="Торпедо",command=lambda:[pyperclip.copy(TorpedoLF),infoteam()])
TorpedoLFK.place(x=790, y=174)
TorpedoLFK['state'] = 'disabled'
ss75='СШ-75 - 2'
ssh75=Button(window,font=Style, text="СШ-75 - 2",command=lambda:[pyperclip.copy(ss75),infoteam()])
ssh75.place(x=790, y=199)
ssh75['state'] = 'disabled'
SpartaLFK='Спартак-2'
SpartakLFK=Button(window, font=Style,text="'Спартак-2",command=lambda:[pyperclip.copy(SpartaLFK),infoteam()])
SpartakLFK.place(x=790, y=224)
SpartakLFK['state'] = 'disabled'
TrudovRezrvLFK='Трудовые резервы-2'
TrudovRezrviLFK=Button(window,font=Style, text="Трудовые резервы-2",command=lambda:[pyperclip.copy(TrudovRezrvLFK),infoteam()])
TrudovRezrviLFK.place(x=790, y=249)
SokolLF='Сокол'
SokolLFK=Button(window,font=Style, text="Сокол",command=lambda:[pyperclip.copy(SokolLF),infoteam()])
SokolLFK.place(x=790, y=274)
RodinaLF='Родина-м'
RodinaLFK=Button(window, font=Style,text="Родина-м",command=lambda:[pyperclip.copy(RodinaLF),infoteam()])
RodinaLFK.place(x=790, y=299)
RodinaLFK['state'] = 'disabled'
Krasnogvard='Красногвардеец'
KrasnogvardLFK=Button(window, font=Style,text="Красногвардеец",command=lambda:[pyperclip.copy(Krasnogvard),infoteam()])
KrasnogvardLFK.place(x=790, y=324)
KuncevLFK='Кунцево'
KuncevoLFK=Button(window,font=Style, text="Кунцево",command=lambda:[pyperclip.copy(KuncevLFK),infoteam()])
KuncevoLFK.place(x=790, y=349)

# СОЗДАНИЕ ФАЙЛА КОМАНДЫ
creatTEAM = Entry(window,width=17)  
creatTEAM.place(x=1, y=550)
CREATTEAM = Button(window,text="Создание файла Команды",command=creatfileteam) 
CREATTEAM.place(x=110, y=550)
# ВНЕСЕНИЕ ССЫЛОК В ФАЙЛ ШКОЛЫ
addlinkschool1 = Entry(window,width=17)  
addlinkschool1.place(x=1, y=580)
linkschool=Label(window,text='Файл команды')
linkschool.place(x=110, y=580)
addlinkschool2 = Entry(window,width=17)  
addlinkschool2.place(x=1, y=610)
linkfile=Label(window,text='Папка')
linkfile.place(x=110, y=610)
ADDLINKSCHOOL = Button(window,text="Внести ссылки в файл Школы",command=addlinksforschool) 
ADDLINKSCHOOL.place(x=1, y=640)
ADDLINKSCHOOL = Button(window,text="Внести ссылки в файл ЛФК",command=addlinksforLFK) 
ADDLINKSCHOOL.place(x=1, y=670)
#КДК
KDK2=Button(window,text="КДК",command=KDK)
KDK2.place(x=900, y=0)
#НОВОСТИ
webbrowser.register('Chrome', None, webbrowser.BackgroundBrowser('C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe'))
RfsLink='https://rfs.ru'
def RFSNEW():
    def NovostiRFS1():
        webbrowser.get(using='Chrome').open_new_tab(RfsLink+Rfslink1)
    def NovostiRFS2():
        webbrowser.get(using='Chrome').open_new_tab(RfsLink+Rfslink2)
    def NovostiRFS3():
        webbrowser.get(using='Chrome').open_new_tab(RfsLink+Rfslink3)
    def NovostiRFS4():
        webbrowser.get(using='Chrome').open_new_tab(RfsLink+Rfslink4)
    def NovostiRFS5():
        webbrowser.get(using='Chrome').open_new_tab(RfsLink+Rfslink5)
    def NovostiRFS6():
        webbrowser.get(using='Chrome').open_new_tab(RfsLink+Rfslink6)
    def NovostiRFS7():
        webbrowser.get(using='Chrome').open_new_tab(RfsLink+Rfslink7)
    def NovostiRFS8():
        webbrowser.get(using='Chrome').open_new_tab(RfsLink+Rfslink8)
    def NovostiRFS9():
        webbrowser.get(using='Chrome').open_new_tab(RfsLink+Rfslink9)
    def NovostiRFS10():
        webbrowser.get(using='Chrome').open_new_tab(RfsLink+Rfslink10)
    RFSLink='https://rfs.ru/news'
    RFSReq=requests.get(RFSLink)
    RFSReq.raise_for_status()
    RFStree=lxml.html.document_fromstring(RFSReq.text)
    news=RFStree.xpath('//*[@id="p0"]/div[3]/div[1]/*[@class="news-list-article "]/*[@class="news-list-article-cont"]/*[@class="news-list-article-cont-title"]/text()')
    Rfsnew1=news[0]
    Rfsnew2=news[1]
    Rfsnew3=news[2]
    Rfsnew4=news[3]
    Rfsnew5=news[4]
    Rfsnew6=news[5]
    Rfsnew7=news[6]
    Rfsnew8=news[7]
    Rfsnew9=news[8]
    Rfsnew10=news[9]
    Rfsnews1=Button(window,text=Rfsnew1,bg='orange',command=NovostiRFS1)
    Rfsnews1.place(x=1000, y=0)
    Rfsnews2=Button(window,text=Rfsnew2,bg='orange',command=NovostiRFS2)
    Rfsnews2.place(x=1000, y=27)
    Rfsnews3=Button(window,text=Rfsnew3,bg='orange',command=NovostiRFS3)
    Rfsnews3.place(x=1000, y=54)
    Rfsnews4=Button(window,text=Rfsnew4,bg='orange',command=NovostiRFS4)
    Rfsnews4.place(x=1000, y=81)
    Rfsnews5=Button(window,text=Rfsnew5,bg='orange',command=NovostiRFS5)
    Rfsnews5.place(x=1000, y=108)
    Rfsnews6=Button(window,text=Rfsnew6,bg='orange',command=NovostiRFS6)
    Rfsnews6.place(x=1000, y=135)
    Rfsnews7=Button(window,text=Rfsnew7,bg='orange',command=NovostiRFS7)
    Rfsnews7.place(x=1000, y=162)
    Rfsnews8=Button(window,text=Rfsnew8,bg='orange',command=NovostiRFS8)
    Rfsnews8.place(x=1000, y=189)
    Rfsnews9=Button(window,text=Rfsnew9,bg='orange',command=NovostiRFS9)
    Rfsnews9.place(x=1000, y=216)
    Rfsnews10=Button(window,text=Rfsnew10,bg='orange',command=NovostiRFS10)
    Rfsnews10.place(x=1000, y=243)
    links=RFStree.xpath('//*[@id="p0"]/div[3]/div[1]/*[@class="news-list-article "]/*[@class="news-list-article-cont"]/a')
    Rfslink1=links[0].get('href')
    Rfslink2=links[1].get('href')
    Rfslink3=links[2].get('href')
    Rfslink4=links[3].get('href')
    Rfslink5=links[4].get('href')
    Rfslink6=links[5].get('href')
    Rfslink7=links[6].get('href')
    Rfslink8=links[7].get('href')
    Rfslink9=links[8].get('href')
    Rfslink10=links[9].get('href')
    Rfsnews1.after(4000000, RFSNEW)
Tread = threading.Thread(target=RFSNEW,daemon = True)
Tread.start()
#def tick():
 #   label.after(200, tick)
  #  label['text'] = time.strftime('%H:%M:%S')
#root=Tk()
#label = Label(font='sans 20')
#label.pack()
#label.after_idle(tick)
#root.mainloop()
#bg- фон
#fg-цвет текст
#focus- фокус на элементе при старте
window.mainloop()
